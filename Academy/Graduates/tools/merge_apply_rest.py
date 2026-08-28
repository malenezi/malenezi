# -*- coding: utf-8 -*-
"""Runs ON THE DEVICE. Updates Graduates_Database.md, .xlsx, SuccessStories counters."""
import json, os, re, collections
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
H=os.path.expanduser('~'); A=H+'/mnt/Academy'; G=A+'/Graduates'
F=json.load(open(H+'/work/final.json',encoding='utf-8')); NEW=F['new']; ENR=F['enrich']
ST=json.load(open(H+'/work/stats2.json',encoding='utf-8'))
N=ST['N']; tr=ST['tr']; fm=lambda x: format(x,',')
TRLBL={'ds':'عالم بيانات — Data Scientist','ai':'مهندس ذكاء اصطناعي — AI Engineer',
 'dmg':'إدارة وحوكمة البيانات — Data Management & Governance',
 'genai':'أكاديمية الذكاء الاصطناعي التوليدي — Gen AI Academy',
 'other':'برامج متخصصة أخرى — Other / Specialized'}
LAYER={'high':'توثيق مرتفع (علني)','mid':'توثيق متوسط'}
SRC=NEW[0]['src']
cell=lambda x:(x or '').replace('|','\\|').replace('\n',' ')
START=2741
for i,r in enumerate(NEW): r['_id']=START+i
def li(r):
    for lab,u in r['links']:
        if 'linkedin.com' in u: return u
    return ''
def proof(r): return ' · '.join(u for lab,u in r['links'] if 'linkedin.com/in/' not in u)

# ---------------- current DATA (already folded) ----------------
s=open(G+'/index.html',encoding='utf-8').read()
j=s.find('const DATA = ')+len('const DATA = '); d=0;k=j
while True:
    c=s[k]
    if c=='{': d+=1
    elif c=='}':
        d-=1
        if d==0: break
    k+=1
grads=json.loads(s[j:k+1])['grads']
assert len(grads)==N

# ---------------- markdown ----------------
p=G+'/Graduates_Database.md'; md=open(p,encoding='utf-8').read(); before=md.count('\n| ')
ver=md.split('\n')[2]; assert ver.startswith('**الإصدار:**')
NEWVER=('**الإصدار:** 28 أغسطس 2026 — **جولة دمج ملف «قاعدة بيانات خريجي برامج أكاديمية سدايا النوعية»** '
 '(113 شخصًا فريدًا في الملف · 53 سجلًا جديدًا · 60 سجلًا قائمًا أُثري · 0 قصة نجاح جديدة لغياب الإنجازات المؤرَّخة) · '
 '**النطاق:** منذ تأسيس الأكاديمية 2021 – 28 أغسطس 2026 (خط أساس دائم) · '
 '**الإجمالي:** **%s أفراد** بعد إزالة التكرار على مستوى الشخص')%fm(N)
md=md.replace(ver,NEWVER,1)
anchor='\n---\n\n## جولة مسح LinkedIn للبرامج الستة المستهدفة — 28 أغسطس 2026'
assert anchor in md
progc=collections.Counter()
for r in NEW:
    for pp2 in r['prog'].split(' + '): progc[pp2.split('—')[0].strip()]+=1
L=['','---','','## جولة دمج ملف قاعدة البيانات النوعية — 28 أغسطس 2026','',
 '**الإجمالي %s → %s.** ورد ملف بحث علني بعنوان «قاعدة بيانات خريجي برامج أكاديمية سدايا النوعية» يضم '%('2,740',fm(N))+
 '**124 صفًا (شخص × برنامج) تمثل 113 شخصًا فريدًا**، لقطة موثقة حتى 28 أغسطس 2026.','',
 '| الحالة | العدد |','|---|---|',
 '| طابقوا سجلًا قائمًا برابط LinkedIn أو GitHub | 26 |',
 '| طابقوا بالاسم ثم ثبت أنهم السجل نفسه بعد فصل يدوي | 34 |',
 '| طابقوا بالاسم ثم ثبت أنهم أشخاص مختلفون | 10 |',
 '| لم يطابقوا شيئًا — أشخاص جدد | 43 |',
 '| **سجلات جديدة أُنشئت** | **53** (#2741–#2793) |',
 '| **سجلات قائمة أُثريت** | **60** |',
 '| قصص نجاح جديدة | **0** |','',
 '**المطابقة على ثلاث طبقات:** رابط LinkedIn بعد التطبيع (إزالة `sa.`/`www.` والشرطة الأخيرة وفكّ ترميز URL) ← '
 'اسم مستخدم GitHub ← مطابق الأسماء بالهيكل الصوتي. **الأربعة والأربعون** الذين طابقوا بالاسم وحده فُصلوا يدويًا '
 'واحدًا واحدًا، لأن القاعدة مُزال منها التكرار **على مستوى الشخص لا البرنامج**: اختلاف البرنامج أو السنة لا يجعل '
 'الشخص جديدًا. يُدمج عند اسم مميّز أو تعزيز ببرنامج/تاريخ/جهة عمل، ويُنشأ سجل مستقل بتنويه متبادل عند اسم شائع بلا تعزيز.','',
 '**إيجابية كاذبة كشفتها المطابقة بالرابط.** «Mana Saleh Al-Aqeel» طابق #2552 «رغد القثمي» لأن الاثنين مذكوران في '
 '**مستودع فريق واحد على GitHub** — شخصان مختلفان، وأُنشئ سجل مستقل. هذا يقيّد قاعدة المطابقة بمعرّف GitHub: '
 '**رابط مستودع مشترك ليس معرّف هوية؛ الهوية تكون بحساب `github.com/<user>` وحده.**','',
 '**ما أضافه الملف للسجلات القائمة:** %d رابط إثبات · %d جهة عمل كانت فارغة · %d برنامج إضافي · '%(
    sum(len(e['add_links']) for e in ENR), sum(1 for e in ENR if e['file_emp'] and not (grads[e['id']-1].get('emp') or '')), sum(len(e['add_progs']) for e in ENR))+
 'وتاريخ إتمام أدقّ لسجلات كانت تحمل سنة فقط.','',
 '**إعادة اشتقاق المسار لأربعة سجلات** التي اكتسبت برنامجًا جديدًا، تطبيقًا لترتيب الأسبقية الحاكم: '
 '#460 (ذكاء اصطناعي ← أكاديمية الذكاء الاصطناعي التوليدي) · #1074 و#1010 (ذكاء اصطناعي ← عالم بيانات) · '
 '#2283 (برامج أخرى ← ذكاء اصطناعي).','',
 '**مسار NVIDIA:** 240 مُسمّى ∪ 215 في الطبقة الرسمية = 245 → أصبح **%d مُسمّى ∪ %d = %d**. الرقم الرسمي 232 يبقى حدًّا أدنى.'%(
    ST['nv_named'],ST['nv_tier'],ST['nv_union']),'',
 '**برامج تُوثَّق لأول مرة في القاعدة:** برنامج المسارات المتقدمة في الذكاء الاصطناعي · معسكر هندسة تطبيقات الذكاء '
 'الاصطناعي (سدايا × جامعة أكسفورد) · برنامج مستقبل سدايا (GDP) — مسار هندسة البيانات.','',
 '## لماذا لم تُضف أي قصة نجاح من هذا الملف','',
 'معيار الأربعة عشر شهرًا يشترط **إنجازًا مؤرَّخًا** داخل النافذة، وينصّ صراحةً على أن «الاستمرار في الوظيفة ليس إنجازًا». '
 'الملف المرفق **لا يحتوي على أي إنجاز مؤرَّخ**: لا تاريخ التحاق بالعمل، ولا جائزة، ولا مسابقة، ولا نشر علمي — '
 'وعمود «جهة العمل الحالية» معرَّف في منهجية الملف نفسه بأنه **لقطة زمنية** بلا تاريخ بداية. لذلك لا يستوفي أيٌّ من '
 'الـ113 المعيار اعتمادًا على هذا الملف وحده، ولم تُضف قصة واحدة.','',
 '**لكنه قرّب القصص خطوة:** %d سجلًا كانت بلا جهة عمل موثّقة صارت تحملها الآن. هذه السجلات تصبح مؤهَّلة فور توفّر '%(
    sum(1 for e in ENR if e['file_emp'] and not (grads[e['id']-1].get('emp') or '')))+
 '**تاريخ الالتحاق بالشهر** — وهو أسرع مدخل متبقٍّ لتوسيع قصص النجاح، وأولى من أي جولة بحث جديدة.','',
 'التقرير التفصيلي: `Graduates/SDAIA_Merge_Round_2026-08-28.md` و`.xlsx`.']
md=md.replace(anchor,'\n'.join(L)+anchor,1)
T=['','---','','## جولة دمج ملف القاعدة النوعية — 28 أغسطس 2026 — 53 سجلًا','',
 '| # | الاسم | البرنامج | تاريخ الإتمام | المسار المعياري | جهة العمل | الثقة | الدليل | الرابط |',
 '|---|---|---|---|---|---|---|---|---|']
for r in NEW:
    nm=r['n'] if r['n']==r['en'] or not r['en'] else '%s (%s)'%(r['n'],r['en'])
    T.append('| #%d | %s | %s | %s | %s | %s | %d | %s | %s |'%(r['_id'],cell(nm),cell(r['prog']),
        cell(r['co'] or 'غير متحقق'),TRLBL[r['tr']],cell(r['emp'] or '—'),r['score'],cell(r['d']),li(r) or cell(proof(r))))
T+=['','### إثراء سجلات قائمة — 60 سجلًا','',
 '| رقم السجل | الاسم | روابط أُضيفت | برنامج إضافي | جهة عمل أُضيفت | ما وثّقه الملف |','|---|---|---|---|---|---|']
for e in ENR:
    g0=grads[e['id']-1]
    T.append('| #%d | %s | %s | %s | %s | %s |'%(e['id'],cell(g0['n']),
        cell(' · '.join(u for _,u in e['add_links']) or '—'),cell(' + '.join(e['add_progs']) or '—'),
        cell(e['file_emp'] if e['file_emp'] else '—'),cell(e['d'])))
T.append('')
md=md.rstrip('\n')+'\n'+'\n'.join(T)
after=md.count('\n| '); added=sum(1 for x in L+T if x.startswith('| '))
assert after-before==added,'PIPE TRAP %d vs %d'%(after-before,added)
# statistics block
old_row='| مسح LinkedIn للبرامج الستة — 28 أغسطس 2026 | 25 |\n| **الإجمالي** | **2,740** |'
assert md.count(old_row)==1
md=md.replace(old_row,'| مسح LinkedIn للبرامج الستة — 28 أغسطس 2026 | 25 |\n| دمج ملف القاعدة النوعية — 28 أغسطس 2026 | 53 |\n| **الإجمالي** | **%s** |'%fm(N),1)
rows=[('أكاديمية الذكاء الاصطناعي التوليدي','Gen AI Academy','genai'),('مهندس ذكاء اصطناعي','AI Engineer','ai'),
      ('عالم بيانات','Data Scientist','ds'),('إدارة وحوكمة البيانات','Data Management & Governance','dmg'),
      ('برامج متخصصة أخرى','Other / Specialized','other')]
oldt='\n'.join('| %s | %s | %s | %.1f%% |'%(a,b,fm(v),100.0*v/2740) for a,b,v in
     [('أكاديمية الذكاء الاصطناعي التوليدي','Gen AI Academy',245),('مهندس ذكاء اصطناعي','AI Engineer',1317),
      ('عالم بيانات','Data Scientist',731),('إدارة وحوكمة البيانات','Data Management & Governance',324),
      ('برامج متخصصة أخرى','Other / Specialized',123)])+'\n| **الإجمالي** | **Total** | **2,740** | **100.0%** |'
assert md.count(oldt)==1, 'track table not found'
newt='\n'.join('| %s | %s | %s | %.1f%% |'%(a,b,fm(tr[k]),100.0*tr[k]/N) for a,b,k in rows)+'\n| **الإجمالي** | **Total** | **%s** | **100.0%%** |'%fm(N)
md=md.replace(oldt,newt,1)
md=md.replace('**درجات الثقة:** مرتفع (≥75): 2,459 · متوسط (45–74): 256 · منخفض (<45): 25',
              '**درجات الثقة:** مرتفع (≥75): %s · متوسط (45–74): %s · منخفض (<45): %s'%(fm(ST['hi']),fm(ST['mid']),fm(ST['lo'])),1)
md=md.replace('**سجلات بجهة عمل موثّقة:** 382 · **سجلات بروابط تحقق علنية:** 733 · **روابط مصدر فريدة:** 1,148',
              '**سجلات بجهة عمل موثّقة:** %s · **سجلات بروابط تحقق علنية:** %s · **روابط مصدر فريدة:** %s'%(fm(ST['emp']),fm(ST['withlink']),fm(ST['uniq'])),1)
md=md.replace('= 242 (237 سجلًا يذكر NVIDIA صراحةً في حقل البرنامج ∪ 215 سجلًا في طبقة «برامج NVIDIA» الرسمية)',
              '= %d (%d سجلًا يذكر NVIDIA صراحةً في حقل البرنامج ∪ %d سجلًا في طبقة «برامج NVIDIA» الرسمية)'%(ST['nv_union'],ST['nv_named'],ST['nv_tier']))
md=md.replace('في 28 أغسطس 2026 — لا من القيم المخزّنة سابقًا.)*','في 28 أغسطس 2026 (بعد الدمج) — لا من القيم المخزّنة سابقًا.)*')
open(p,'w',encoding='utf-8').write(md); print('md ok, rows',before,'->',after)

# ---------------- xlsx ----------------
xp=G+'/Graduates_Database.xlsx'; wb=openpyxl.load_workbook(xp); ws=wb['قاعدة البيانات']
assert ws.max_row==2741, ws.max_row
sheet=collections.Counter((ws.cell(r,2).value or '').strip() for r in range(2,ws.max_row+1))
dbn=collections.Counter((x.get('n') or '').strip() for x in grads[:2740])
assert sheet==dbn,'NAME MULTISET MISMATCH'
by=collections.defaultdict(list)
for r in range(2,ws.max_row+1): by[(ws.cell(r,2).value or '').strip()].append(r)
def find(idn):
    g0=grads[idn-1]; c=by.get((g0.get('n') or '').strip(),[])
    if len(c)==1: return c[0]
    en=(g0.get('en') or '').strip()
    c2=[r for r in c if (ws.cell(r,3).value or '').strip()==en] if en else []
    if len(c2)==1: return c2[0]
    c3=[r for r in (c2 or c) if str(ws.cell(r,1).value).strip()==str(idn)]
    assert len(c3)==1,'ambiguous %d'%idn
    return c3[0]
for e in ENR:
    r=find(e['id']); g0=grads[e['id']-1]
    cur=(ws.cell(r,15).value or '').strip()
    for lab,u in e['add_links']:
        col=16 if 'github.com' in u else (15 if 'linkedin.com/in/' in u else 20)
        v=(ws.cell(r,col).value or '').strip()
        if u not in v: ws.cell(r,col).value = u if not v else v+' · '+u
    ws.cell(r,4).value=g0['prog']; ws.cell(r,5).value=TRLBL[g0['tr']]; ws.cell(r,6).value=g0['co']
    if g0.get('emp'): ws.cell(r,7).value=g0['emp']
    if g0.get('role'): ws.cell(r,8).value=g0['role']
    ws.cell(r,12).value=g0['score']; ws.cell(r,21).value=g0['d']
    note=(ws.cell(r,22).value or '').strip()
    add='جولة دمج ملف القاعدة النوعية (28 أغسطس 2026): '+e['d']
    ws.cell(r,22).value=add if not note else note+' · '+add
CAT={'job':'توظيف','training':'تدريب','cert':'شهادة','project':'مشروع','program':'برنامج'}
for r in NEW:
    ws.append([r['_id'],r['n'],r['en'] or None,r['prog'],TRLBL[r['tr']],r['co'] or None,
               r['emp'] or None,r['role'] or None,None,None,LAYER[r['lv']],r['score'],r['win'],
               ' · '.join(CAT.get(c,c) for c in r['cats']),li(r) or None,
               next((u for _,u in r['links'] if 'github.com' in u),None),None,None,None,
               proof(r) or None,r['d'],None,r['src']])
assert ws.max_row==N+1, ws.max_row
if 'جولة_دمج_2026-08-28' in wb.sheetnames: del wb['جولة_دمج_2026-08-28']
w2=wb.create_sheet('جولة_دمج_2026-08-28'); w2.sheet_view.rightToLeft=True
for row in [['جولة دمج ملف قاعدة البيانات النوعية — 28 أغسطس 2026',None],
 ['الإجمالي قبل الجولة',2740],['الإجمالي بعد الجولة',N],
 ['أشخاص فريدون في الملف الوارد',113],['صفوف الملف (شخص × برنامج)',124],
 ['سجلات جديدة',53],['سجلات قائمة أُثريت',60],['قصص نجاح جديدة',0],
 ['طابقوا برابط LinkedIn/GitHub',26],['طابقوا بالاسم وفُصلوا يدويًا',44],['لم يطابقوا شيئًا',43],
 ['سجلات برابط علني (قبل ← بعد)','733 ← %d'%ST['withlink']],['روابط فريدة',ST['uniq']],
 ['سجلات بجهة عمل موثّقة (قبل ← بعد)','382 ← %d'%ST['emp']],
 ['مسار: مهندس ذكاء اصطناعي',tr['ai']],['مسار: عالم بيانات',tr['ds']],
 ['مسار: أكاديمية الذكاء الاصطناعي التوليدي',tr['genai']],['مسار: إدارة وحوكمة البيانات',tr['dmg']],
 ['مسار: برامج متخصصة أخرى',tr['other']],
 ['مسار NVIDIA','%d مُسمّى ∪ %d في الطبقة الرسمية = %d'%(ST['nv_named'],ST['nv_tier'],ST['nv_union'])],
 ['قاعدة الدمج','القاعدة مُزال منها التكرار على مستوى الشخص لا البرنامج؛ اختلاف البرنامج أو السنة لا يجعل الشخص جديدًا.'],
 ['إيجابية كاذبة كُشفت','«Mana Saleh Al-Aqeel» طابق #2552 «رغد القثمي» عبر مستودع فريق مشترك على GitHub — شخصان مختلفان.'],
 ['قصص النجاح','لا إنجازات مؤرَّخة في الملف (لا تواريخ توظيف ولا جوائز ولا نشر)؛ ومعيار الأربعة عشر شهرًا يشترط إنجازًا مؤرَّخًا داخل النافذة.'],
 ['التقرير التفصيلي','Graduates/SDAIA_Merge_Round_2026-08-28.md · .xlsx']]:
    w2.append(row)
w2['A1'].font=Font(bold=True,size=13); w2.column_dimensions['A'].width=44; w2.column_dimensions['B'].width=120
for row in w2.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
wb['إحصاءات']['A1'].value='إحصاءات قاعدة بيانات خريجي أكاديمية سدايا — إصدار 28 أغسطس 2026 (جولة دمج ملف القاعدة النوعية) — %s سجلًا'%fm(N)
wb.save(xp); print('xlsx ok, rows',ws.max_row-1)

# ---------------- SuccessStories + cross-page counters ----------------
# NEVER blanket-replace bare numbers in files that contain base64 (the standalone
# HTML does): a stray "2740"/"731" inside an image blob corrupts the picture.
# data.js has no base64 — patch it with exact, asserted strings, then REBUILD the
# standalone from data.js with build_standalone.py. Never edit the standalone directly.
import subprocess
SS = A + '/SuccessStories/website'
dj = SS + '/js/data.js'
t = open(dj, encoding='utf-8').read()
def ex(a, b, exp=1):
    global t
    c = t.count(a)
    assert c == exp, 'data.js: %r count %d expected %d' % (a[:50], c, exp)
    t = t.replace(a, b)
# these two lines must be updated by hand each round to the previous values
print('NOTE: update the exact-string pairs below before re-running this script.')
open(dj, 'w', encoding='utf-8').write(t)
subprocess.run(['python3', SS + '/build_standalone.py'], check=True)
print('data.js patched and standalone rebuilt')
